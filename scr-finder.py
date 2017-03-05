import openpyxl
import ipaddress
import sys

FILENAME = 'full.xlsx'
FULL_ANALYSIS = True

def expand_range(iprangestr):
    if is_ip_range(iprangestr):
        split = iprangestr.split('-')
        start_ip = split[0]
        iprange = (iprangestr.split('.')[-1]).split('-')
        iprange = list(range(int(iprange[0]),int(iprange[1])+1))
        netobj = ipaddress.IPv4Network(start_ip+'/24', strict=False)
        iprange = ([str(netobj[i]) for i in iprange])
        return iprange

def is_ip_range(teststr):
    return True if '-' in teststr else False

def split_colon_separated_ips(ipstring):
    if ';' in ipstring:
        splits = ipstring.split(';')
        ip_start = '.'.join(splits[0].split('.')[0:3])
        for i in range(1,len(splits)):
            splits[i] = ip_start +'.'+splits[i]
        for item in splits:
            if is_ip_range(item):
                splits += expand_range(item)
        final_list = list()
        for item in splits:
            if not is_ip_range(item):
                final_list.append(item)
        return final_list

if __name__ == '__main__':

    print("Please enter name of output file (without .xlsx): ")
    filename = 'test'

    print ("PHASE 1: Loading Excel Workbook %s" % FILENAME)
    wb = openpyxl.load_workbook(FILENAME)
    sheet = wb.get_sheet_by_name('TSSPM Report')

    #Checking against matchnets sheet
    s2 = wb.get_sheet_by_name('matchnets')

    # Set up the outpoot Workbook
    outxlsx = openpyxl.Workbook()
    outsheet = outxlsx.active
    outsheet.title = 'Outage-Affected-Traffic'
    # Write the header
    outsheet['A1'] = 'NAR#'
    outsheet['B1'] = 'Requester'
    outsheet['C1'] = 'Requester E-Mail'
    outsheet['D1'] = 'Application Name'
    outsheet['E1'] = 'Application Owner'
    outsheet['F1'] = 'Application Owner E-Mail'
    outsheet['G1'] = 'Department Name'
    outsheet['H1'] = 'Networks Matched'
    outsheet['I1'] = 'Affected IP Addresses'


    nar_dict = dict()
    row = 3
    matched=0
    print ("PHASE 2: Loading information about NARs")
    while sheet.cell(row=row, column=1).value!=None:
        matched_networks = set()
        matched_ips_in_nar = set()
        nar = sheet.cell(row=row, column=1).value
        sources = []
        destinations = []
        for col in range(53,62):
            value =  sheet.cell(row=row, column=col).value
            if value != None:
                value = value.replace("\n",',')
                value = value.replace(',,',',')
                if value != '':
                    destinations += (value.split(','))
        for col in range(53,62):
            value =  sheet.cell(row=row, column=col).value
            if value != None:
                value = value.replace("\n",',')
                value = value.replace(',,',',')
                if value != '':
                    sources += (value.split(','))
        #Expand ranges and multi-ip definitions
        destinations =  list(set(destinations))
        expanded_dsts = list()
        for dst in destinations:
            if ';' in dst:
                expanded_dsts += split_colon_separated_ips(dst)
                for i, edst in enumerate(expanded_dsts):
                    if '-' in edst:
                        expanded_edsts = expand_range(dst)
                        expanded_dsts.remove(edst)
                        expanded_dsts += expanded_edsts
            elif '-' in dst:
                expanded_dsts += expand_range(dst)
            elif dst!='*':
                expanded_dsts.append(dst)
        expanded_srcs = list()
        for src in sources:
            if ';' in src:
                expanded_srcs += split_colon_separated_ips(src)
            elif '-' in src:
                expanded_srcs += expand_range(src)
            elif src!='*':
                expanded_srcs.append(src)
        nar_dict[nar] = [expanded_srcs, expanded_dsts, len(expanded_srcs), len(expanded_dsts)]
        r2 = 1
        flag = False
        while s2.cell(row=r2, column=1).value!=None:
            sys.stdout.write('\rProcessing %s --- %d/%d NARS MATCHED/TESTED' % (nar, matched, row-2) )
            network = (s2.cell(row=r2,column=1).value).strip()
            for ip in nar_dict[nar][0]+nar_dict[nar][1]:
                if ip.endswith('.'):
                    ip = ip[:-1]
                if '/' in ip:
                    ipnet = ip
                elif ip!='':
                    ipnet = ip+'/32'
                try:
                    c1 = ipaddress.ip_network(ipnet, strict=False).overlaps(ipaddress.IPv4Network(network))
                    c2 = ipaddress.ip_network(ipnet, strict=False).prefixlen >= ipaddress.IPv4Network(network).prefixlen
                except:
                    print("ERROR - NAR: %s, IP: %s, Network: %s" %(nar,ip,network))
                    pass
                if c1 and c2:
                    if flag == False:
                        matched += 1
                    flag = True
                    matched_networks.add(network)
                    matched_ips_in_nar.add(ip)
                    #print("\n%s - IP %s is contained in %s" % (nar, ipnet, network))
                if FULL_ANALYSIS == False:
                    break
            r2 += 1
        if flag==True:
            """
            1 outsheet['A1'] = 'NAR#'
            2 outsheet['B1'] = 'Requester'
            3 outsheet['C1'] = 'Requester E-Mail'
            4 outsheet['D1'] = 'Application Name'
            5 outsheet['E1'] = 'Application Owner'
            6 outsheet['F1'] = 'Application Owner E-Mail'
            7 outsheet['G1'] = 'Department Name'
            8 outsheet['H1'] = 'Networks Matched'
            9 outsheet['I1'] = 'Affected IP Addresses'
            """

            outsheet.cell(row=matched+2, column=1).value = nar
            outsheet.cell(row=matched+2, column=2).value = sheet.cell(row=row, column=11).value
            outsheet.cell(row=matched+2, column=3).value = sheet.cell(row=row, column=4).value
            outsheet.cell(row=matched+2, column=4).value = sheet.cell(row=row, column=16).value
            outsheet.cell(row=matched+2, column=5).value = sheet.cell(row=row, column=27).value
            outsheet.cell(row=matched+2, column=6).value = sheet.cell(row=row, column=28).value
            outsheet.cell(row=matched+2, column=7).value = sheet.cell(row=row, column=16).value
            outsheet.cell(row=matched+2, column=8).value = str(matched_networks)
            outsheet.cell(row=matched+2, column=9).value = str(matched_ips_in_nar)
        sys.stdout.write('\rProcessing %s --- %d/%d NARS MATCHED/TESTED' % (nar, matched, row - 2))
        row+= 1
    print('\n----OPERATION COMPLETED----\n')
    outxlsx.save(filename + '.xlsx')
